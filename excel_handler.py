# excel_handler.py 只管 openpyxl 的所有操作，包含完善的文件防占用机制。
import time

import openpyxl

import config

# 文件被占用时的提示回调，默认走命令行 input；GUI 可注入弹窗提示
_retry_prompt = input


def set_retry_prompt(func):
    global _retry_prompt
    _retry_prompt = func


def _prompt_retry(message):
    return _retry_prompt(message)


def update_excel(data):
    print("\n[EXCEL] 正在打开文件...")
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    ws = wb.active

    new_row = [data.get(col_name, "") for col_name in config.COLUMNS]
    ws.append(new_row)
    print(f"[EXCEL] 已写入第 {ws.max_row} 行")

    while True:
        try:
            wb.save(config.EXCEL_PATH)
            wb.close()
            print("[OK] Excel 主表已更新")
            break
        except PermissionError:
            _prompt_retry("[WARN] Excel 主表被占用，请关闭后按回车重试...")


def _find_progress_record(company, job_name, stage):
    if not config.EXCEL_PATH.exists():
        return -1, False

    try:
        while True:
            try:
                wb = openpyxl.load_workbook(config.EXCEL_PATH)
                break
            except PermissionError:
                print("[WARN] Excel 文件被占用，等待重试...")
                time.sleep(1)
    except Exception:
        return -1, False

    if config.PROGRESS_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return -1, False

    ws = wb[config.PROGRESS_SHEET_NAME]
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row_idx, 3).value == company
            and ws.cell(row_idx, 4).value == job_name
            and ws.cell(row_idx, 5).value == stage
        ):
            wb.close()
            return row_idx, True

    wb.close()
    return -1, False


def update_progress_sheet(data, is_new=True):
    company = data.get("公司", "").strip()
    job_name = data.get("投递志愿与顺序", "").strip()
    stage = data.get("当前进度", "").strip()
    date_val = data.get("对应日期", "").strip()

    if not all([company, job_name, stage, date_val]):
        print("[INFO] 四项核心信息不完整，跳过附表同步。")
        return

    row_idx, found = _find_progress_record(company, job_name, stage)
    try:
        wb = openpyxl.load_workbook(config.EXCEL_PATH)
        ws = wb[config.PROGRESS_SHEET_NAME]

        if found:
            ws.cell(row_idx, 6).value = date_val
            print(f"[UPDATE] 附表已更新记录: {company} - {job_name} - {stage}")
        else:
            new_row = ws.max_row + 1
            event_title = (
                f"新增投递：{company}-{job_name}"
                if is_new
                else f"进度变更：{company}-{job_name} → {stage}"
            )
            row_data = [event_title, "", company, job_name, stage, date_val, ""]
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(new_row, col_idx).value = val
            print(f"[NEW] 附表新增记录: {company} - {job_name} - {stage}")

        while True:
            try:
                wb.save(config.EXCEL_PATH)
                wb.close()
                print("[OK] 附表已保存。")
                break
            except PermissionError:
                _prompt_retry("[WARN] Excel 附表被占用，请关闭后按回车重试...")
    except Exception as e:
        print(f"[ERROR] 保存附表失败: {e}")


def is_duplicate_record(data):
    """
    检查主表是否已存在相同的记录（公司 + base + 岗位）。
    """
    company = data.get("公司", "").strip()
    base = data.get("base", "").strip()
    job = data.get("投递志愿与顺序", "").strip()

    # 如果核心信息为空，保守起见不判定为重复
    if not company or not job:
        return False

    if not config.EXCEL_PATH.exists():
        return False

    try:
        while True:
            try:
                wb = openpyxl.load_workbook(config.EXCEL_PATH)
                break
            except PermissionError:
                print("[WARN] 查重时 Excel 被占用，等待重试...")
                import time

                time.sleep(1)
    except Exception:
        return False

    ws = wb.active

    # 动态获取对应的列号 (索引从 1 开始)
    try:
        col_company = config.COLUMNS.index("公司") + 1
        col_base = config.COLUMNS.index("base") + 1
        col_job = config.COLUMNS.index("投递志愿与顺序") + 1
    except ValueError:
        wb.close()
        return False

    # 遍历主表比对（从第2行开始，跳过表头）
    for row_idx in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row_idx, col_company).value or "").strip()
        row_base = str(ws.cell(row_idx, col_base).value or "").strip()
        row_job = str(ws.cell(row_idx, col_job).value or "").strip()

        # 三要素完全匹配
        if row_company == company and row_base == base and row_job == job:
            wb.close()
            return True

    wb.close()
    return False
